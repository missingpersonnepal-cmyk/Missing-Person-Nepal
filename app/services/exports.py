from __future__ import annotations

from io import BytesIO, StringIO
import csv

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from ..models import Disaster, MissingPerson

PERSON_HEADERS = [
    "Case ID",
    "Name",
    "Nepali Name",
    "Age",
    "Gender",
    "Last Seen Date",
    "Last Seen Time",
    "Last Seen Location",
    "Clothing",
    "Identification Details",
    "Public Contact",
    "Residential Address (Admin Only)",
    "Private/Family Contact (Admin Only)",
    "Source Count",
    "Published",
    "Archived",
]


def _people(db: Session, disaster_id: int) -> list[MissingPerson]:
    return list(
        db.scalars(
            select(MissingPerson)
            .options(selectinload(MissingPerson.sources))
            .where(MissingPerson.disaster_id == disaster_id)
            .order_by(MissingPerson.case_number)
        ).all()
    )


def _row(person: MissingPerson) -> list[object]:
    return [
        person.case_number,
        person.name,
        person.name_ne or "",
        person.age if person.age is not None else "",
        person.gender or "",
        person.last_seen_date.isoformat() if person.last_seen_date else "",
        person.last_seen_time.isoformat(timespec="minutes") if person.last_seen_time else "",
        person.last_seen_location,
        person.clothing or "",
        person.identification_details or "",
        person.public_contact_number or "",
        person.residential_address_private or "",
        person.private_contact_number or "",
        len(person.sources),
        person.published,
        person.archived,
    ]


def build_csv(db: Session, disaster_id: int) -> bytes:
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(PERSON_HEADERS)
    for person in _people(db, disaster_id):
        writer.writerow(_row(person))
    return output.getvalue().encode("utf-8-sig")


def build_xlsx(db: Session, disaster_id: int) -> bytes:
    disaster = db.get(Disaster, disaster_id)
    if disaster is None:
        raise ValueError("Unknown disaster")

    wb = Workbook()
    ws = wb.active
    ws.title = "Missing People"
    ws.append(PERSON_HEADERS)
    people = _people(db, disaster_id)
    for person in people:
        ws.append(_row(person))

    source_ws = wb.create_sheet("Sources")
    source_ws.append(["Case ID", "Platform", "Source URL", "Source Name", "Discovered At"])
    for person in people:
        for source in person.sources:
            source_ws.append([
                person.case_number,
                source.platform,
                source.url,
                source.source_name or "",
                source.discovered_at.isoformat() if source.discovered_at else "",
            ])

    meta_ws = wb.create_sheet("Metadata")
    meta_ws.append(["Disaster", disaster.name])
    meta_ws.append(["Event Code", disaster.code])
    meta_ws.append(["Start Date", disaster.start_date.isoformat()])
    meta_ws.append(["Total Persons", len(people)])
    meta_ws.append(["Total Sources", sum(len(p.sources) for p in people)])

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
